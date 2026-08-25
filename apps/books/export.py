"""Excel va PDF hisobotlari.

openpyxl va reportlab bu yerda ATAYLAB fayl boshida chaqirilmagan.

Sabab tezlik. `urls.py` shu modulni har safar o'qiydi, ya'ni ilgari
server har ko'tarilganda ikkala kutubxona ham yuklanardi — bu qo'shimcha
~170 ms vaqt va ~27 MB xotira. Render'ning bepul tarifida protsessor
o'ndan bir (0.1 CPU) bo'lgani uchun o'sha 170 ms bir necha sekundga
cho'ziladi va foydalanuvchi buni "sayt uyg'onmayapti" deb ko'radi.

Hisobot esa kuniga bir-ikki marta yuklanadi. Shuning uchun kutubxonalar
faqat tugma bosilganda — funksiya ichida — yuklanadi.
"""

from django.contrib.auth.decorators import login_required
from django.http import HttpResponse

from .models import Author, Book, Purchase


def _excel():
    """openpyxl'ni shu yerda yuklaydi va kerakli nomlarni qaytaradi."""
    from openpyxl import Workbook
    from openpyxl.styles import Font

    return Workbook, Font


def _pdf():
    """reportlab'ni shu yerda yuklaydi va kerakli nomlarni qaytaradi."""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import cm
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    return SimpleDocTemplate, A4, cm, getSampleStyleSheet, Paragraph, Spacer, Table, TableStyle, colors


def _books_queryset(user):
    qs = Book.objects.select_related("author", "genre", "seller")
    if user.is_staff or user.is_superuser:
        return qs
    return qs.filter(seller=user)


def _authors_queryset(user):
    if user.is_staff or user.is_superuser:
        return Author.objects.all()
    return Author.objects.filter(books__seller=user).distinct()


@login_required
def export_books_excel(request):
    Workbook, Font = _excel()
    books = _books_queryset(request.user)

    wb = Workbook()
    ws = wb.active
    ws.title = "Kitoblar"

    headers = ["#", "Nomi", "Muallif", "Janr", "Til", "Sahifa", "Narx", "Reyting", "Sotuvchi", "Qo'shilgan sana"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for i, book in enumerate(books, start=1):
        ws.append(
            [
                i,
                book.title,
                book.author.full_name,
                book.genre.name if book.genre else "-",
                book.get_language_display(),
                book.pages,
                float(book.price),
                book.average_rating,
                book.seller.username,
                book.created_at.strftime("%Y-%m-%d"),
            ]
        )

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="kitoblar.xlsx"'
    wb.save(response)
    return response


@login_required
def export_authors_excel(request):
    Workbook, Font = _excel()
    authors = _authors_queryset(request.user)

    wb = Workbook()
    ws = wb.active
    ws.title = "Mualliflar"

    headers = ["#", "Ism familiya", "Tug'ilgan sana", "Kitoblar soni"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for i, author in enumerate(authors, start=1):
        ws.append([i, author.full_name, str(author.birth_date) if author.birth_date else "-", author.books_count])

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="mualliflar.xlsx"'
    wb.save(response)
    return response


@login_required
def export_books_pdf(request):
    (SimpleDocTemplate, A4, cm, getSampleStyleSheet, Paragraph, Spacer,
     Table, TableStyle, colors) = _pdf()
    books = _books_queryset(request.user)

    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="kitoblar.pdf"'

    doc = SimpleDocTemplate(response, pagesize=A4, topMargin=1.5 * cm, bottomMargin=1.5 * cm)
    styles = getSampleStyleSheet()
    elements = [Paragraph("Kitoblar ro'yxati", styles["Title"]), Spacer(1, 12)]

    data = [["#", "Nomi", "Muallif", "Til", "Narx", "Reyting"]]
    for i, book in enumerate(books, start=1):
        data.append([i, book.title, book.author.full_name, book.get_language_display(), f"{book.price}", book.average_rating])

    table = Table(data, repeatRows=1)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2c3e50")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f2f2f2")]),
                ("FONTSIZE", (0, 0), (-1, -1), 9),
            ]
        )
    )
    elements.append(table)
    doc.build(elements)
    return response


@login_required
def export_authors_pdf(request):
    (SimpleDocTemplate, A4, cm, getSampleStyleSheet, Paragraph, Spacer,
     Table, TableStyle, colors) = _pdf()
    authors = _authors_queryset(request.user)

    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="mualliflar.pdf"'

    doc = SimpleDocTemplate(response, pagesize=A4, topMargin=1.5 * cm, bottomMargin=1.5 * cm)
    styles = getSampleStyleSheet()
    elements = [Paragraph("Mualliflar ro'yxati", styles["Title"]), Spacer(1, 12)]

    data = [["#", "Ism familiya", "Tug'ilgan sana", "Kitoblar soni"]]
    for i, author in enumerate(authors, start=1):
        data.append([i, author.full_name, str(author.birth_date) if author.birth_date else "-", author.books_count])

    table = Table(data, repeatRows=1)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2c3e50")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f2f2f2")]),
                ("FONTSIZE", (0, 0), (-1, -1), 9),
            ]
        )
    )
    elements.append(table)
    doc.build(elements)
    return response


@login_required
def export_sales_excel(request):
    """Sotuvchining savdo hisoboti.

    Administrator uchun - barcha savdolar, sotuvchi uchun - faqat o'ziniki.
    """
    Workbook, Font = _excel()
    sales = Purchase.objects.select_related("book", "book__author", "buyer")
    if not (request.user.is_staff or request.user.is_superuser):
        sales = sales.filter(book__seller=request.user)

    wb = Workbook()
    ws = wb.active
    ws.title = "Savdo"

    headers = ["#", "Sana", "Kitob", "Muallif", "Xaridor", "Summa (so'm)", "Karta"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    total = 0
    for i, sale in enumerate(sales, start=1):
        total += float(sale.price_paid)
        ws.append(
            [
                i,
                sale.purchased_at.strftime("%Y-%m-%d %H:%M"),
                sale.book.title,
                sale.book.author.full_name,
                sale.buyer.username,
                float(sale.price_paid),
                f"**** {sale.card_last4}" if sale.card_last4 else "-",
            ]
        )

    ws.append([])
    summary = ["", "", "", "", "JAMI", total, ""]
    ws.append(summary)
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True)

    for column, width in zip("ABCDEFG", (5, 18, 34, 26, 18, 16, 14)):
        ws.column_dimensions[column].width = width

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="savdo-hisoboti.xlsx"'
    wb.save(response)
    return response
